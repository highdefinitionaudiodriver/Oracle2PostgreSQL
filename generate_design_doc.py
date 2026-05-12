"""
Oracle2PostgreSQL - Design Document Generator
Generates design_document.xlsx with 5 sheets.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Shared styles ──
FONT_HEADER = Font(name="Arial", bold=True, size=11, color="FFFFFF")
FONT_SECTION = Font(name="Arial", bold=True, size=11, color="1A5276")
FONT_NORMAL = Font(name="Arial", size=10)
FONT_CODE = Font(name="Consolas", size=9)
FONT_TITLE = Font(name="Arial", bold=True, size=14, color="1A5276")

FILL_HEADER = PatternFill("solid", fgColor="2E86C1")
FILL_ALT = PatternFill("solid", fgColor="EBF5FB")
FILL_SECTION = PatternFill("solid", fgColor="D6EAF8")
FILL_AUTO = PatternFill("solid", fgColor="D5F5E3")
FILL_REVIEW = PatternFill("solid", fgColor="FDEBD0")
FILL_MANUAL = PatternFill("solid", fgColor="FADBD8")

ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_TL = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header_row(ws, row, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_C
        cell.border = THIN_BORDER


def style_data_row(ws, row, col_count, alt=False):
    for c in range(1, col_count + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_L
        cell.border = THIN_BORDER
        if alt:
            cell.fill = FILL_ALT


def add_title(ws, row, title, col_span=8):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_TITLE
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)


def add_section(ws, row, title, col_span=8):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_SECTION
    cell.fill = FILL_SECTION
    cell.border = THIN_BORDER
    for c in range(2, col_span + 1):
        ws.cell(row=row, column=c).fill = FILL_SECTION
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=col_span)


# ====================================================================
# Sheet 1: 機能一覧表
# ====================================================================
def create_function_list(wb):
    ws = wb.active
    ws.title = "1.機能一覧表"
    cols = ["機能ID", "機能カテゴリ", "機能名", "概要", "対象ユーザー", "入力", "出力", "備考"]
    widths = [10, 16, 24, 50, 16, 24, 24, 30]

    add_title(ws, 1, "Oracle2PostgreSQL v1.0.0 - 機能一覧表（基本設計）")
    r = 3
    for i, h in enumerate(cols, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(cols))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    data = [
        # ── コア変換機能 ──
        ["F-001", "コア変換", "Oracle SQL パース", "Oracle SQL/PL-SQL ファイルを抽象構文木 (AST) にパースする。CREATE TABLE, INDEX, VIEW, SEQUENCE, SYNONYM, PROCEDURE, FUNCTION, PACKAGE, TRIGGER の 30 種類の文型を認識。", "開発者 / DBA", "Oracle SQL ファイル (.sql, .pls 等)", "OracleSchema (AST)", "文字列リテラル内の誤検出を防止する構造解析"],
        ["F-002", "コア変換", "データ型変換", "Oracle 固有のデータ型を PostgreSQL 互換型に変換する (23 ルール)。VARCHAR2→VARCHAR, NUMBER→NUMERIC, DATE→TIMESTAMP, CLOB→TEXT, BLOB→BYTEA 等。", "開発者 / DBA", "OracleSchema (AST)", "変換済み SQL + ChangeRecord", "精度・スケール付きの NUMBER にも対応"],
        ["F-003", "コア変換", "関数変換", "Oracle 固有関数を PostgreSQL 関数にマッピングする (26 ルール)。SYSDATE→CURRENT_TIMESTAMP, NVL→COALESCE, SYS_GUID→gen_random_uuid 等。", "開発者 / DBA", "OracleSchema (AST)", "変換済み SQL + ChangeRecord", "DECODE は CASE 変換ガイド (REVIEW)"],
        ["F-004", "コア変換", "構文変換", "Oracle 固有構文を PostgreSQL 標準に変換する (18 ルール)。FROM DUAL 削除, MINUS→EXCEPT, ヒント削除, STORAGE 句削除等。", "開発者 / DBA", "OracleSchema (AST)", "変換済み SQL + ChangeRecord", "CONNECT BY は MANUAL 判定"],
        ["F-005", "コア変換", "PL/SQL → PL/pgSQL 変換", "PL/SQL プロシージャ・関数をPL/pgSQL 形式に変換する (21 ルール)。IS→AS $$, DBMS_OUTPUT→RAISE NOTICE, RAISE_APPLICATION_ERROR→RAISE EXCEPTION 等。", "開発者 / DBA", "OracleSchema (AST)", "変換済み PL/pgSQL", "パラメータ方向 (IN/OUT/IN OUT) を保持"],
        ["F-006", "コア変換", "トリガー変換", "Oracle トリガーを PostgreSQL の「トリガー関数 + CREATE TRIGGER」形式に分離変換する (3 ルール)。:NEW/:OLD の自動変換を含む。", "開発者 / DBA", "OracleSchema (AST)", "トリガー関数 + CREATE TRIGGER", "RETURN NEW/OLD の自動挿入"],
        ["F-007", "コア変換", "シーケンス変換", "Oracle シーケンス構文を PostgreSQL 形式に変換する (4 ルール)。NOCACHE/NOORDER 削除, NOCYCLE→NO CYCLE。", "開発者 / DBA", "OracleSchema (AST)", "変換済み CREATE SEQUENCE", "seq.NEXTVAL→nextval('seq') も対応"],
        ["F-008", "コア変換", "シノニム変換", "Oracle SYNONYM を PostgreSQL VIEW に変換する。代替として search_path 設定もガイドする。", "開発者 / DBA", "OracleSchema (AST)", "CREATE OR REPLACE VIEW", "PUBLIC SYNONYM 対応"],
        ["F-009", "コア変換", "パッケージ変換ガイド", "Oracle PACKAGE を Schema + 個別関数へ変換するガイドコメントを自動挿入する。パッケージは構造が複雑なため MANUAL 判定。", "開発者 / DBA", "OracleSchema (AST)", "ガイドコメント付き SQL", "変数は設定テーブルへの移行を推奨"],
        ["F-010", "コア変換", "ストレージ句除去", "Oracle 固有の STORAGE, TABLESPACE, PCTFREE, INITRANS, LOGGING 等の物理格納句を自動除去する。", "開発者 / DBA", "OracleSchema (AST)", "クリーンな DDL", "インデックス・テーブル両方に適用"],
        # ── 出力・レポート ──
        ["F-011", "出力", "PostgreSQL ファイル生成", "変換結果を _pg.sql ファイルとして出力する。ヘッダーコメント (ソースファイル名, 変更サマリー) を付与。", "開発者 / DBA", "TransformResult", "*_pg.sql ファイル", "出力エンコーディング設定可能"],
        ["F-012", "出力", "HTML 移行レポート生成", "全ファイルの変更を集計し、ブラウザ閲覧可能な HTML レポートを生成する。Executive Summary, カテゴリ別集計, REVIEW/MANUAL 一覧, ファイル別詳細を含む。", "開発者 / DBA / PM", "List[TransformResult]", "migration_report_*.html", "多言語対応 (日本語/英語等)"],
        ["F-013", "出力", "CSV 移行レポート生成", "全ファイルの変更を CSV として出力する。Excel/スプレッドシートでの分析・フィルタリングに対応。", "開発者 / DBA / PM", "List[TransformResult]", "migration_report_*.csv", "UTF-8 BOM 付き (Excel 対応)"],
        ["F-014", "出力", "バックアップ作成", "変換前の原本ファイルをタイムスタンプ付きバックアップディレクトリにコピーする。", "開発者 / DBA", "原本 SQL ファイル", "_backup/YYYYMMDD_HHMMSS/", "設定で無効化可能"],
        # ── UI ──
        ["F-015", "UI", "GUI モード (Tkinter)", "入出力フォルダ選択、設定ファイル読込、変換オプション ON/OFF、プログレスバー、リアルタイムログ、キャンセル機能を持つ GUI。", "開発者 / DBA", "マウス/キーボード操作", "GUI ウィンドウ", "バックグラウンドスレッド実行"],
        ["F-016", "UI", "CLI モード", "コマンドライン引数で入出力ディレクトリ、エンコーディング、変換オプションを指定して実行する CUI モード。CI/CD パイプラインに組み込み可能。", "開発者 / DevOps", "コマンドライン引数", "コンソール出力 + ファイル生成", "-i, -o, -c, --lang 等14オプション"],
        ["F-017", "UI", "62 言語国際化 (i18n)", "GUI ラベル、ボタン、ログメッセージ、レポートを 62 言語で表示する。言語は GUI のコンボボックスまたは --lang オプションで切り替え。", "全ユーザー", "言語コード (ja, en 等)", "翻訳済み UI テキスト", "39 の翻訳キーを管理"],
        # ── 設定・運用 ──
        ["F-018", "設定", "YAML 設定ファイル", "config.yaml でルールカテゴリ ON/OFF、個別ルール無効化、スキーママッピング、エンコーディング、ログ設定を外部設定化する。", "開発者 / DBA", "config.yaml", "AppConfig データクラス", "CLI 引数で上書き可能"],
        ["F-019", "設定", "スキーママッピング", "config.yaml で Oracle スキーマ名 → PostgreSQL スキーマ名のマッピングを定義する (例: HR → public)。", "DBA", "config.yaml", "変換時のスキーマ名置換", "複数マッピング定義可能"],
        ["F-020", "運用", "構造化ロギング", "コンソール (INFO) + RotatingFileHandler (DEBUG) のデュアル出力ロガー。ファイル変換・ルール適用・エラーを構造化フォーマットで記録。", "開発者 / 運用担当", "変換処理イベント", "logs/migration.log", "5MB/3世代ローテーション"],
        ["F-021", "運用", "Docker 対応", "Dockerfile + docker-compose.yml で CLI 専用コンテナ + PostgreSQL 検証コンテナを提供する。", "DevOps", "docker compose build/up", "変換済みファイル + レポート", "マルチステージビルド"],
        ["F-022", "運用", "CI/CD パイプライン", "GitHub Actions で Push/PR 時に pytest (Python 3.11/3.12 マトリックス) + Docker ビルドを自動実行する。", "開発者 / DevOps", "git push / PR", "テスト結果 + ビルド結果", ".github/workflows/ci.yml"],
        ["F-023", "品質", "テストスイート", "39 件のユニットテスト + 統合テスト (pytest)。パーサー、ルール、トランスフォーマー、ジェネレーター、レポート、E2E を網羅。", "開発者", "pytest 実行", "テスト結果", "カバレッジ計測対応"],
    ]

    for i, row_data in enumerate(data):
        r = 4 + i
        for j, val in enumerate(row_data):
            ws.cell(row=r, column=j + 1, value=val)
        style_data_row(ws, r, len(cols), alt=(i % 2 == 1))

    ws.sheet_properties.pageSetUpPr = None
    ws.freeze_panes = "A4"


# ====================================================================
# Sheet 2: API仕様書（CLI / Internal API）
# ====================================================================
def create_api_spec(wb):
    ws = wb.create_sheet("2.API仕様書")
    cols = ["API ID", "インタフェース種別", "エンドポイント / メソッド", "パラメータ", "データ型", "必須", "デフォルト値", "処理概要", "戻り値 / レスポンス"]
    widths = [10, 14, 34, 22, 14, 8, 16, 44, 30]

    add_title(ws, 1, "Oracle2PostgreSQL v1.0.0 - API仕様書（詳細設計）")
    r = 3
    for i, h in enumerate(cols, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(cols))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    data = [
        # ── CLI ──
        ["", "", "", "", "", "", "", "", ""],
        ["CLI-001", "CLI 引数", "-i, --input", "<directory_path>", "str", "Yes*", "-", "変換対象の Oracle SQL ファイルが格納されたディレクトリを指定する。再帰的にファイルを探索する。", "N/A (入力パラメータ)"],
        ["CLI-002", "CLI 引数", "-o, --output", "<directory_path>", "str", "Yes*", "-", "変換結果の PostgreSQL SQL ファイルとレポートを出力するディレクトリを指定する。存在しない場合は自動作成。", "N/A (入力パラメータ)"],
        ["CLI-003", "CLI 引数", "-c, --config", "<file_path>", "str", "No", "config.yaml", "YAML 設定ファイルのパスを指定する。ルール ON/OFF、スキーママッピング、ログ設定等を外部制御する。", "N/A (入力パラメータ)"],
        ["CLI-004", "CLI 引数", "-e, --encoding", "<encoding_name>", "str", "No", "utf-8 (config)", "入力ファイルの文字エンコーディングを指定する。config.yaml の値を上書きする。", "N/A (入力パラメータ)"],
        ["CLI-005", "CLI 引数", "--extensions", "<ext1,ext2,...>", "str", "No", ".sql,.pls,...(config)", "変換対象のファイル拡張子をカンマ区切りで指定する。config.yaml の値を上書きする。", "N/A (入力パラメータ)"],
        ["CLI-006", "CLI 引数", "--lang", "<lang_code>", "str", "No", "ja (config)", "UI / レポートの表示言語コードを指定する (62言語対応)。", "N/A (入力パラメータ)"],
        ["CLI-007", "CLI フラグ", "--no-datatypes", "(フラグ)", "bool", "No", "False", "データ型変換 (DATATYPE カテゴリ) をスキップする。", "N/A (入力パラメータ)"],
        ["CLI-008", "CLI フラグ", "--no-plsql", "(フラグ)", "bool", "No", "False", "PL/SQL → PL/pgSQL 変換 (PLSQL カテゴリ) をスキップする。", "N/A (入力パラメータ)"],
        ["CLI-009", "CLI フラグ", "--no-sequences", "(フラグ)", "bool", "No", "False", "シーケンス変換 (SEQUENCE カテゴリ) をスキップする。", "N/A (入力パラメータ)"],
        ["CLI-010", "CLI フラグ", "--no-synonyms", "(フラグ)", "bool", "No", "False", "シノニム変換 (SYNONYM カテゴリ) をスキップする。", "N/A (入力パラメータ)"],
        ["CLI-011", "CLI フラグ", "--no-packages", "(フラグ)", "bool", "No", "False", "パッケージ変換をスキップする。", "N/A (入力パラメータ)"],
        ["CLI-012", "CLI フラグ", "--no-triggers", "(フラグ)", "bool", "No", "False", "トリガー変換 (TRIGGER カテゴリ) をスキップする。", "N/A (入力パラメータ)"],
        ["CLI-013", "CLI フラグ", "--no-report", "(フラグ)", "bool", "No", "False", "HTML/CSV レポート生成をスキップする。", "N/A (入力パラメータ)"],
        ["CLI-014", "CLI フラグ", "--no-backup", "(フラグ)", "bool", "No", "False", "原本ファイルのバックアップ作成をスキップする。", "N/A (入力パラメータ)"],
        ["", "", "", "", "", "", "", "", ""],
        # ── Internal Python API ──
        ["", "", "", "", "", "", "", "", ""],
        ["API-001", "Internal API", "OracleParser.parse_file(filepath)", "filepath: str", "str", "Yes", "-", "Oracle SQL ファイルを読み込み、AST (OracleSchema) にパースする。エンコーディング設定に従いファイルを読み込む。", "OracleSchema"],
        ["API-002", "Internal API", "OracleParser.parse_string(content)", "content: str", "str", "Yes", "-", "Oracle SQL 文字列を AST (OracleSchema) にパースする。ファイルI/Oを介さず直接文字列を解析する。", "OracleSchema"],
        ["API-003", "Internal API", "PostgresTransformer.transform(schema)", "schema: OracleSchema", "OracleSchema", "Yes", "-", "OracleSchema に 104 の変換ルールを適用し、PostgreSQL 互換の SQL に変換する。各変更を ChangeRecord として記録する。", "TransformResult"],
        ["API-004", "Internal API", "PostgresCodeGenerator.generate(result)", "result: TransformResult", "TransformResult", "Yes", "-", "変換結果をファイルに書き出す。ヘッダーコメント、変更サマリーを付与する。", "str (出力ファイルパス)"],
        ["API-005", "Internal API", "ReportGenerator.generate(results)", "results: List[TransformResult]", "List", "Yes", "-", "全ファイルの変換結果を集計し、HTML および CSV レポートを生成する。", "Dict[str, str] (html/csv パス)"],
        ["API-006", "Internal API", "BackupManager.backup(filepath)", "filepath: str", "str", "Yes", "-", "指定ファイルをタイムスタンプ付きバックアップディレクトリにコピーする。", "Optional[str] (バックアップパス)"],
        ["API-007", "Internal API", "load_config(config_path)", "config_path: str", "str", "Yes", "-", "YAML ファイルを読み込み、型安全な AppConfig データクラスに変換する。", "AppConfig"],
        ["API-008", "Internal API", "setup_logger(name, ...)", "name, log_file, console_level, file_level, max_bytes, backup_count", "各種", "name のみ", "上記参照", "コンソール + RotatingFileHandler のデュアルロガーを構成する。", "logging.Logger"],
        ["API-009", "Internal API", "I18n.t(key, **kwargs)", "key: str, **kwargs", "str", "Yes", "-", "指定キーの翻訳文字列を返す。kwargs でプレースホルダーを置換する。", "str (翻訳済み文字列)"],
        ["API-010", "Internal API", "MigrationRules.get_rules_by_category(cat)", "category: str", "str", "Yes", "-", "指定カテゴリの全ルールをリストで返す。", "List[MigrationRule]"],
        ["API-011", "Internal API", "AppConfig.to_transform_options()", "(self)", "-", "-", "-", "AppConfig を TransformOptions に変換する。CLI 引数による上書き前の基本オプションを生成する。", "TransformOptions"],
    ]

    # セクション行の挿入
    r = 4
    add_section(ws, r, "CLI インタフェース（コマンドライン引数）", len(cols))
    r += 1
    for i, row_data in enumerate(data):
        if row_data[0] == "":
            if row_data[1] == "":
                if i > 0 and i < len(data) - 1:
                    add_section(ws, r, "Internal Python API（内部関数インタフェース）", len(cols))
                    r += 1
                continue
            continue
        for j, val in enumerate(row_data):
            ws.cell(row=r, column=j + 1, value=val)
        style_data_row(ws, r, len(cols), alt=(i % 2 == 1))
        r += 1

    ws.freeze_panes = "A4"


# ====================================================================
# Sheet 3: テーブル定義書（内部データ構造）
# ====================================================================
def create_table_definitions(wb):
    ws = wb.create_sheet("3.テーブル定義書")
    cols = ["データ構造名", "フィールド名", "Python 型", "論理名", "必須", "デフォルト値", "制約 / 備考"]
    widths = [22, 22, 20, 28, 8, 18, 44]

    add_title(ws, 1, "Oracle2PostgreSQL v1.0.0 - テーブル定義書（詳細設計）")
    ws.cell(row=2, column=1, value="※ 本ツールはDBを使用しないため、内部データ構造（データクラス）の定義を記載する。").font = Font(name="Arial", size=9, italic=True, color="666666")

    r = 4
    for i, h in enumerate(cols, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(cols))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    structures = [
        ("OracleSchema\n(AST ルート)", [
            ["filename", "str", "入力ファイルパス", "No", '""', "パース対象のファイルパス"],
            ["statements", "List[OracleStatement]", "全SQL文リスト", "No", "[]", "パースされた全ステートメント"],
            ["tables", "List[TableDef]", "テーブル定義リスト", "No", "[]", "CREATE TABLE から抽出"],
            ["indexes", "List[IndexDef]", "インデックス定義リスト", "No", "[]", "CREATE INDEX から抽出"],
            ["sequences", "List[SequenceDef]", "シーケンス定義リスト", "No", "[]", "CREATE SEQUENCE から抽出"],
            ["views", "List[ViewDef]", "ビュー定義リスト", "No", "[]", "CREATE VIEW から抽出"],
            ["synonyms", "List[SynonymDef]", "シノニム定義リスト", "No", "[]", "CREATE SYNONYM から抽出"],
            ["procedures", "List[ProcedureDef]", "プロシージャ/関数リスト", "No", "[]", "CREATE PROCEDURE/FUNCTION から抽出"],
            ["packages", "List[PackageDef]", "パッケージ定義リスト", "No", "[]", "CREATE PACKAGE から抽出"],
            ["triggers", "List[TriggerDef]", "トリガー定義リスト", "No", "[]", "CREATE TRIGGER から抽出"],
            ["errors", "List[str]", "パースエラー一覧", "No", "[]", "パース中に発生したエラーメッセージ"],
            ["has_plsql", "bool", "PL/SQL 使用フラグ", "No", "False", "PL/SQL ブロックが存在するか"],
            ["has_oracle_hints", "bool", "ヒント使用フラグ", "No", "False", "/*+ ... */ ヒントが存在するか"],
            ["has_connect_by", "bool", "階層クエリフラグ", "No", "False", "CONNECT BY が存在するか"],
            ["has_outer_join_plus", "bool", "(+) 結合フラグ", "No", "False", "(+) 外部結合が存在するか"],
            ["oracle_functions_used", "List[str]", "使用Oracle関数一覧", "No", "[]", "SYSDATE, NVL 等の検出済み関数名"],
        ]),
        ("TableDef", [
            ["name", "str", "テーブル名", "Yes", "-", "CREATE TABLE で定義されたテーブル名"],
            ["schema", "Optional[str]", "スキーマ名", "No", "None", "schema.table の schema 部分"],
            ["columns", "List[ColumnDef]", "カラム定義リスト", "No", "[]", "テーブル内の全カラム定義"],
            ["constraints", "List[ConstraintDef]", "制約定義リスト", "No", "[]", "PK, FK, UNIQUE, CHECK"],
            ["tablespace", "Optional[str]", "テーブルスペース名", "No", "None", "TABLESPACE 句の値 (変換時に除去)"],
            ["is_temporary", "bool", "一時テーブルフラグ", "No", "False", "GLOBAL TEMPORARY TABLE か否か"],
        ]),
        ("ColumnDef", [
            ["name", "str", "カラム名", "Yes", "-", "カラム識別子"],
            ["data_type", "str", "データ型", "Yes", "-", "VARCHAR2, NUMBER 等"],
            ["precision", "Optional[str]", "精度", "No", "None", "NUMBER(10) の 10"],
            ["scale", "Optional[str]", "スケール", "No", "None", "NUMBER(12,2) の 2"],
            ["nullable", "bool", "NULL 許容", "No", "True", "NOT NULL 指定がなければ True"],
            ["default_value", "Optional[str]", "デフォルト値", "No", "None", "DEFAULT 句の値"],
            ["is_primary_key", "bool", "主キーフラグ", "No", "False", "インラインPRIMARY KEY指定"],
        ]),
        ("ConstraintDef", [
            ["name", "Optional[str]", "制約名", "No", "None", "CONSTRAINT 句の名前"],
            ["constraint_type", "str", "制約種別", "Yes", "-", "PRIMARY KEY / FOREIGN KEY / UNIQUE / CHECK"],
            ["columns", "List[str]", "対象カラム", "No", "[]", "制約が適用されるカラム名リスト"],
            ["ref_table", "Optional[str]", "参照先テーブル", "No", "None", "FK の REFERENCES 先テーブル"],
            ["ref_columns", "List[str]", "参照先カラム", "No", "[]", "FK の REFERENCES 先カラム"],
            ["condition", "Optional[str]", "CHECK 条件式", "No", "None", "CHECK 制約の条件式"],
        ]),
        ("MigrationRule", [
            ["rule_id", "str", "ルールID", "Yes", "-", "DT_001, FN_003 等。カテゴリ接頭辞+番号"],
            ["category", "str", "カテゴリ", "Yes", "-", "DATATYPE / FUNCTION / SYNTAX / PLSQL / SEQUENCE / OBJECT / TRIGGER / SYNONYM"],
            ["severity", "str", "重要度", "Yes", "-", "AUTO / REVIEW / MANUAL"],
            ["old_pattern", "str", "検索パターン (正規表現)", "Yes", "-", "Oracle 側のマッチパターン"],
            ["new_pattern", "str", "置換パターン", "Yes", "-", "PostgreSQL 側の置換文字列"],
            ["description", "str", "英語説明", "Yes", "-", "ルールの概要 (英語)"],
            ["description_ja", "str", "日本語説明", "No", '""', "ルールの概要 (日本語)"],
            ["auto_fix", "bool", "自動適用フラグ", "No", "True", "True=自動置換, False=検出のみ"],
            ["severity_override", "str", "重要度オーバーライド", "No", '""', "auto_fix=False 時の代替重要度"],
        ]),
        ("TransformResult", [
            ["filename", "str", "ソースファイル名", "Yes", "-", "変換元ファイルのパス"],
            ["original_text", "str", "変換前テキスト", "Yes", "-", "元の Oracle SQL テキスト全体"],
            ["transformed_text", "str", "変換後テキスト", "Yes", "-", "PostgreSQL 変換済みテキスト全体"],
            ["changes", "List[ChangeRecord]", "変更記録リスト", "No", "[]", "適用された全変更の詳細記録"],
            ["auto_converted", "int", "自動変換数", "No", "0", "AUTO 判定の変更数"],
            ["needs_review", "int", "要確認数", "No", "0", "REVIEW 判定の変更数"],
            ["manual_only", "int", "手動対応数", "No", "0", "MANUAL 判定の変更数"],
            ["errors", "List[str]", "エラー一覧", "No", "[]", "変換中に発生したエラー"],
        ]),
        ("AppConfig", [
            ["input_encoding", "str", "入力エンコーディング", "No", '"utf-8"', "ソースファイルの文字コード"],
            ["output_encoding", "str", "出力エンコーディング", "No", '"utf-8"', "出力ファイルの文字コード"],
            ["file_extensions", "List[str]", "対象拡張子", "No", "[.sql,.pls,...]", "変換対象のファイル拡張子一覧"],
            ["output_suffix", "str", "出力サフィックス", "No", '"_pg"', "出力ファイル名に付与する接尾辞"],
            ["language", "str", "言語コード", "No", '"ja"', "UI/レポートの言語"],
            ["category_toggles", "Dict[str,bool]", "カテゴリON/OFF", "No", "全True", "ルールカテゴリごとの有効/無効"],
            ["disabled_rules", "Set[str]", "無効化ルール集合", "No", "空集合", "個別に無効化するルールID"],
            ["schema_mapping", "Dict[str,str]", "スキーママッピング", "No", "{}", "Oracle→PostgreSQLスキーマ名変換"],
            ["logging", "LoggingConfig", "ログ設定", "No", "デフォルト", "ログレベル、ファイルパス、ローテーション"],
            ["create_backup", "bool", "バックアップ作成", "No", "True", "原本ファイルのバックアップ有無"],
        ]),
    ]

    r = 5
    for struct_name, fields in structures:
        add_section(ws, r, struct_name, len(cols))
        r += 1
        for i, fld in enumerate(fields):
            ws.cell(row=r, column=1, value=struct_name.split("\n")[0] if i == 0 else "")
            for j, val in enumerate(fld):
                ws.cell(row=r, column=j + 2, value=val)
            style_data_row(ws, r, len(cols), alt=(i % 2 == 1))
            r += 1
        r += 1

    ws.freeze_panes = "A5"


# ====================================================================
# Sheet 4: エラー・ログ定義書
# ====================================================================
def create_error_log_definitions(wb):
    ws = wb.create_sheet("4.エラー・ログ定義書")
    cols = ["ログID", "カテゴリ", "ログレベル", "出力先", "出力タイミング", "メッセージフォーマット", "パラメータ", "対処方法"]
    widths = [12, 14, 10, 14, 28, 44, 24, 36]

    add_title(ws, 1, "Oracle2PostgreSQL v1.0.0 - エラー・ログ定義書（詳細設計）")
    r = 3
    for i, h in enumerate(cols, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(cols))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    data = [
        # ── ライフサイクル ──
        ["LOG-001", "ライフサイクル", "INFO", "コンソール+ファイル", "変換処理開始時", "Migration started: {file_count} files", "file_count: int", "-"],
        ["LOG-002", "ライフサイクル", "INFO", "コンソール+ファイル", "変換処理完了時", "Migration complete: {total} files | Auto:{auto} Review:{review} Manual:{manual}", "total, auto, review, manual: int", "-"],
        # ── ファイル処理 ──
        ["LOG-003", "ファイル処理", "INFO", "コンソール+ファイル", "ファイルパース開始時", "[PARSE] {filename}", "filename: str", "-"],
        ["LOG-004", "ファイル処理", "INFO", "コンソール+ファイル", "ファイル変換開始時", "[TRANSFORM] {filename}", "filename: str", "-"],
        ["LOG-005", "ファイル処理", "INFO", "コンソール+ファイル", "ファイル出力完了時", "[GENERATE] {filename} -> {output_path}", "filename, output_path: str", "-"],
        ["LOG-006", "ファイル処理", "INFO", "コンソール+ファイル", "ファイル変換結果出力時", "[RESULT] {filename}: Auto={auto} Review={review} Manual={manual}", "filename: str, auto/review/manual: int", "-"],
        # ── ルール適用 ──
        ["LOG-007", "ルール適用", "DEBUG", "ファイルのみ", "ルール適用成功時", "[RULE] [{severity}] {rule_id} ({category}): {description} | {filename}:{line}", "rule_id, category, severity, description, filename: str, line: int", "-"],
        ["LOG-008", "ルール適用", "DEBUG", "ファイルのみ", "ルールスキップ時", "[SKIP] {rule_id}: {reason}", "rule_id, reason: str", "config.yaml の disabled_rules 確認"],
        # ── パース警告 ──
        ["LOG-009", "パース", "WARNING", "コンソール+ファイル", "パース中に非致命的エラー検出時", "[PARSE-WARN] {filename}:{line} - {message}", "filename: str, line: int, message: str", "対象の SQL 文を手動確認"],
        ["LOG-010", "パース", "ERROR", "コンソール+ファイル", "パース中に致命的エラー発生時", "[PARSE-ERR] {filename}:{line} - {message}", "filename: str, line: int, message: str", "SQL 構文を修正して再実行"],
        ["LOG-011", "パース", "WARNING", "コンソール+ファイル", "未対応構文検出時", "[SKIP-SYNTAX] {filename}:{line} - Unsupported: {syntax}", "filename: str, line: int, syntax: str", "手動で PostgreSQL 互換に変換"],
        # ── 設定 ──
        ["LOG-012", "設定", "INFO", "コンソール+ファイル", "設定ファイル読み込み成功時", "Config loaded: {config_path}", "config_path: str", "-"],
        ["LOG-013", "設定", "ERROR", "コンソール+ファイル", "設定ファイル読み込み失敗時", "Config error: {message}", "message: str", "config.yaml の構文を確認"],
        ["LOG-014", "設定", "ERROR", "コンソール+ファイル", "入力ディレクトリ不存在時", "Input directory not found: {path}", "path: str", "パスの存在を確認"],
        ["LOG-015", "設定", "ERROR", "コンソール+ファイル", "対象ファイル不存在時", "No matching files found in {path}", "path: str", "拡張子フィルタを確認"],
        # ── 実行時エラー ──
        ["LOG-016", "実行時", "ERROR", "コンソール+ファイル", "ファイル読み込み失敗時", "Failed to read file: {error}", "error: str (OSError/IOError)", "ファイルパス・権限を確認"],
        ["LOG-017", "実行時", "ERROR", "コンソール+ファイル", "変換中の未処理例外発生時", "Unexpected error: {error}", "error: str (Exception)", "スタックトレースをログファイルで確認"],
        # ── 重要度判定 ──
        ["LOG-018", "レポート", "INFO", "HTMLレポート", "AUTO 判定ルール適用時", "[AUTO] [{category}] {rule_id}: {description}", "category, rule_id, description", "変換結果を目視確認 (推奨)"],
        ["LOG-019", "レポート", "WARNING", "HTMLレポート", "REVIEW 判定ルール適用時", "[REVIEW] [{category}] {rule_id}: {description}", "category, rule_id, description", "PostgreSQL 環境で動作検証必須"],
        ["LOG-020", "レポート", "ERROR", "HTMLレポート", "MANUAL 判定ルール検出時", "[MANUAL] [{category}] {rule_id}: {description}", "category, rule_id, description", "手動で PostgreSQL 互換コードに書き換え"],
    ]

    r = 4
    prev_cat = ""
    for i, row_data in enumerate(data):
        cat = row_data[1]
        if cat != prev_cat and i > 0:
            add_section(ws, r, cat, len(cols))
            r += 1
            prev_cat = cat
        elif i == 0:
            add_section(ws, r, cat, len(cols))
            r += 1
            prev_cat = cat
        for j, val in enumerate(row_data):
            ws.cell(row=r, column=j + 1, value=val)
        style_data_row(ws, r, len(cols), alt=(i % 2 == 1))

        lvl = row_data[2]
        if lvl == "ERROR":
            ws.cell(row=r, column=3).fill = FILL_MANUAL
        elif lvl == "WARNING":
            ws.cell(row=r, column=3).fill = FILL_REVIEW
        elif lvl == "DEBUG":
            ws.cell(row=r, column=3).fill = PatternFill("solid", fgColor="E8DAEF")
        r += 1

    ws.freeze_panes = "A4"


# ====================================================================
# Sheet 5: アーキテクチャ図解（Mermaid コード集）
# ====================================================================
def create_architecture_diagrams(wb):
    ws = wb.create_sheet("5.アーキテクチャ図解")
    cols = ["図の種類", "Mermaid コード"]
    widths = [20, 120]

    add_title(ws, 1, "Oracle2PostgreSQL v1.0.0 - アーキテクチャ図解（Mermaid コード集）", col_span=2)
    ws.cell(row=2, column=1, value="※ 下記の Mermaid コードを https://mermaid.live/ 等のレンダラーに貼り付けて図を確認してください。").font = Font(name="Arial", size=9, italic=True, color="666666")

    r = 4
    for i, h in enumerate(cols, 1):
        ws.cell(row=r, column=i, value=h)
    style_header_row(ws, r, len(cols))
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    diagrams = [
        ("システム構成図\n(コンポーネント図)", """graph TB
    subgraph User["ユーザーインタフェース"]
        GUI["GUI Mode<br/>Tkinter"]
        CLI["CLI Mode<br/>argparse"]
        Docker["Docker Mode<br/>docker-compose"]
    end

    subgraph Core["コア変換エンジン"]
        Parser["OracleParser<br/>AST 生成<br/>30 種類の文型認識"]
        Rules["MigrationRules<br/>104 ルール<br/>8 カテゴリ"]
        Transformer["PostgresTransformer<br/>ルール適用<br/>変更記録"]
    end

    subgraph Output["出力モジュール"]
        Generator["PostgresCodeGenerator<br/>_pg.sql 生成"]
        Reporter["ReportGenerator<br/>HTML / CSV"]
        Backup["BackupManager<br/>タイムスタンプ付き"]
    end

    subgraph Config["設定・基盤"]
        YAML["config.yaml<br/>YAML 設定"]
        Loader["ConfigLoader<br/>AppConfig"]
        Logger["Logger<br/>RotatingFileHandler"]
        I18N["I18n<br/>62 言語"]
    end

    GUI --> Parser
    CLI --> Parser
    Docker --> CLI
    Parser --> |OracleSchema| Transformer
    Rules --> Transformer
    Transformer --> |TransformResult| Generator
    Transformer --> |TransformResult| Reporter
    YAML --> Loader
    Loader --> |AppConfig| GUI
    Loader --> |AppConfig| CLI
    Logger --> Core
    I18N --> GUI
    I18N --> Reporter"""),

        ("変換パイプライン\nシーケンス図", """sequenceDiagram
    actor User
    participant Main as main.py
    participant Parser as OracleParser
    participant Trans as PostgresTransformer
    participant Rules as MigrationRules
    participant Gen as PostgresCodeGenerator
    participant Report as ReportGenerator

    User->>Main: 入力フォルダ / 出力フォルダ指定
    Main->>Main: config.yaml 読込 & Logger 初期化

    loop 各 SQL ファイル
        Main->>Parser: parse_file(filepath)
        Parser->>Parser: _split_statements()
        Parser->>Parser: _detect_statement_type()
        Parser->>Parser: _parse_create_table() 等
        Parser-->>Main: OracleSchema (AST)

        Main->>Trans: transform(schema)
        loop 各 Statement
            Trans->>Trans: _transform_statement()
            Trans->>Rules: get_rules_by_category()
            Rules-->>Trans: List[MigrationRule]
            Trans->>Trans: regex 置換 & ChangeRecord 記録
        end
        Trans->>Trans: _apply_generic_rules()
        Trans-->>Main: TransformResult

        Main->>Gen: generate(result)
        Gen-->>Main: 出力ファイルパス
    end

    Main->>Report: generate(all_results)
    Report-->>Main: HTML + CSV パス
    Main-->>User: 完了ログ + レポートパス"""),

        ("データフロー図\n(ER図的構造)", """erDiagram
    OracleSchema ||--o{ OracleStatement : contains
    OracleSchema ||--o{ TableDef : tables
    OracleSchema ||--o{ IndexDef : indexes
    OracleSchema ||--o{ SequenceDef : sequences
    OracleSchema ||--o{ ViewDef : views
    OracleSchema ||--o{ SynonymDef : synonyms
    OracleSchema ||--o{ ProcedureDef : procedures
    OracleSchema ||--o{ PackageDef : packages
    OracleSchema ||--o{ TriggerDef : triggers

    TableDef ||--o{ ColumnDef : columns
    TableDef ||--o{ ConstraintDef : constraints

    ProcedureDef ||--o{ ParameterDef : parameters

    PackageDef ||--o{ ProcedureDef : procedures
    PackageDef ||--o{ ProcedureDef : functions

    OracleStatement ||--o| TableDef : table_def
    OracleStatement ||--o| IndexDef : index_def
    OracleStatement ||--o| SequenceDef : sequence_def
    OracleStatement ||--o| ViewDef : view_def
    OracleStatement ||--o| ProcedureDef : procedure_def
    OracleStatement ||--o| PackageDef : package_def
    OracleStatement ||--o| TriggerDef : trigger_def

    TransformResult ||--o{ ChangeRecord : changes
    MigrationRules ||--o{ MigrationRule : rules"""),

        ("GUI 画面遷移図", """stateDiagram-v2
    [*] --> MainWindow: アプリ起動

    state MainWindow {
        [*] --> Idle: 初期状態

        Idle --> FolderSelect: 参照ボタン押下
        FolderSelect --> Idle: フォルダ選択完了

        Idle --> ConfigLoad: 設定ファイル参照
        ConfigLoad --> Idle: config.yaml 読込完了<br/>チェックボックス自動反映

        Idle --> LangChange: 言語コンボボックス変更
        LangChange --> Idle: UI テキスト切替完了

        Idle --> Converting: 変換ボタン押下
        Converting --> Converting: プログレスバー更新<br/>ログ出力
        Converting --> Idle: 変換完了
        Converting --> Cancelled: キャンセルボタン押下
        Cancelled --> Idle: 処理中断完了
    }

    MainWindow --> [*]: ウィンドウ閉じる"""),

        ("変換ルール\nカテゴリ構成図", """pie title 変換ルール 104件 カテゴリ別内訳
    "FUNCTION (26)" : 26
    "DATATYPE (23)" : 23
    "PLSQL (21)" : 21
    "SYNTAX (18)" : 18
    "OBJECT (8)" : 8
    "SEQUENCE (4)" : 4
    "TRIGGER (3)" : 3
    "SYNONYM (1)" : 1"""),

        ("デプロイメント構成図\n(Docker)", """graph TB
    subgraph Host["ホストマシン"]
        Input["./samples/<br/>Oracle SQL"]
        Output["./output/<br/>PostgreSQL SQL"]
        Config["./config.yaml"]
    end

    subgraph DockerCompose["docker-compose.yml"]
        subgraph MigSvc["migration サービス"]
            MigApp["Oracle2PostgreSQL<br/>Python 3.12-slim<br/>CLI モード"]
        end
        subgraph TestSvc["test サービス"]
            TestApp["pytest<br/>39 テスト実行"]
        end
        subgraph PgSvc["postgres サービス"]
            PgDB["PostgreSQL 16<br/>migration_test DB<br/>変換結果検証用"]
        end
    end

    Input -->|volume mount<br/>readonly| MigApp
    Config -->|volume mount| MigApp
    MigApp -->|volume mount| Output
    MigApp -.->|検証用接続| PgDB
    TestSvc -.->|テスト実行| MigApp

    subgraph CI["GitHub Actions CI"]
        Matrix["Python 3.11 / 3.12<br/>マトリックス"]
        DockerBuild["Docker ビルド<br/>+ コンテナ内テスト"]
    end"""),
    ]

    r = 5
    for diagram_name, mermaid_code in diagrams:
        ws.cell(row=r, column=1, value=diagram_name)
        ws.cell(row=r, column=1).font = FONT_SECTION
        ws.cell(row=r, column=1).alignment = ALIGN_TL
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=mermaid_code.strip())
        ws.cell(row=r, column=2).font = FONT_CODE
        ws.cell(row=r, column=2).alignment = ALIGN_TL
        ws.cell(row=r, column=2).border = THIN_BORDER
        line_count = mermaid_code.strip().count("\n") + 1
        ws.row_dimensions[r].height = max(15 * line_count, 200)
        r += 2

    ws.freeze_panes = "A5"


# ====================================================================
# Main
# ====================================================================
def main():
    wb = Workbook()

    create_function_list(wb)
    create_api_spec(wb)
    create_table_definitions(wb)
    create_error_log_definitions(wb)
    create_architecture_diagrams(wb)

    out = "design_document.xlsx"
    wb.save(out)
    print(f"Generated: {out}")


if __name__ == "__main__":
    main()
